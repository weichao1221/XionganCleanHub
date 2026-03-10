import json
import difflib
import json
import sys
from statistics import mean

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QBrush, QIcon, QDoubleValidator, QFont, QStandardItemModel, QStandardItem
from PySide6.QtGui import QPainter, QTextDocument
from PySide6.QtWidgets import (
    QSplitter, QTreeWidget, QTreeWidgetItem, QTabWidget, QDialog, QSystemTrayIcon, QWidget,
    QVBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QHeaderView, QFileDialog, QMessageBox, QHBoxLayout, QLabel, QLineEdit, QComboBox, QCheckBox, QTreeView,
    QApplication, QSizePolicy
)
from PySide6.QtWidgets import QStyledItemDelegate, QStyleOptionViewItem
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app_paths import resource_path
from utils import functions
import pandas as pd


class qingbiaoResult(QDialog):
    def __init__(self, zongjia_jieguo, fuzhi_jieguo, zero_jieguo, qingdan_jieguo, jiamisuo_jieguo, _10_jieguo, pianchalv, tray_icon):
        super().__init__()
        self.setWindowTitle("清标结果")
        self.zongjia_jieguo = zongjia_jieguo
        self.fuzhi_jieguo = fuzhi_jieguo
        self.zero_jieguo = zero_jieguo
        self.jiamisuo_jieguo = jiamisuo_jieguo
        self._10_jieguo = _10_jieguo
        self.pianchalv = pianchalv
        self.tray_icon = tray_icon
        self.new_result = qingdan_jieguo
        self.init()
        self.resize(900, 900)

    def init(self):
        main_layout = QVBoxLayout()
        self.tab = QTabWidget()

        tab_1 = zongjia_window(self.zongjia_jieguo, self.tray_icon)
        tab_2 = fuzhi_window(self.fuzhi_jieguo, self.tray_icon)
        tab_3 = fuzhi_window(self.zero_jieguo, self.tray_icon)
        tab_7 = qingDanCompareWindow(self.new_result, self.tray_icon)
        tab_5 = LockCheckWindow(self.jiamisuo_jieguo, self.tray_icon)
        tab_6 = PianchaWindow(self._10_jieguo, self.pianchalv, self.tray_icon)

        self.tab.addTab(tab_1, "总价限价结果")
        self.tab.addTab(tab_2, "单价为负值清单")
        self.tab.addTab(tab_3, "单价为0清单")
        self.tab.addTab(tab_7, "清单不一致")
        self.tab.addTab(tab_5, "硬件检测结果")
        self.tab.addTab(tab_6, "价格偏差清单")

        main_layout.addWidget(self.tab)
        self.setLayout(main_layout)


class zongjia_window(QWidget):
    def __init__(self, zongjia_jieguo, tray_icon):
        super().__init__()
        self.zongjia_jieguo = zongjia_jieguo
        self.tray_icon = tray_icon
        self.init()

    def init(self):
        main_layout = QVBoxLayout()
        self.table = QTableWidget()
        # 填充数据
        headers = ["投标人", "投标金额（元）", "控制价金额（元）", "投标下浮率（%）", "是否超限价"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self.zongjia_jieguo))

        for i, item in enumerate(self.zongjia_jieguo):
            for j, value in enumerate(item):
                self.table.setItem(i, j, QTableWidgetItem(str(value)))

        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)  # 禁止编辑

        self.daochubiaoge = QPushButton("导出表格")
        self.daochubiaoge.clicked.connect(self.daochu_table)

        main_layout.addWidget(self.table)
        main_layout.addWidget(self.daochubiaoge)
        self.setLayout(main_layout)

    def daochu_table(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存总价限价结果", "总价限价结果.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "总价限价结果"

        # 写入表头
        headers = ["投标人", "投标金额（元）", "控制价金额（元）", "投标下浮率（%）", "是否超限价"]
        ws.append(headers)

        # 写入数据
        for row in range(self.table.rowCount()):
            row_data = []
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            ws.append(row_data)

        try:
            wb.save(file_path)
            self.tray_icon.showMessage("成功", f"总价限价结果已导出：\n{file_path}")
        except Exception as e:
            self.tray_icon.showMessage("错误", f"导出失败：\n{e}")


class fuzhi_window(QWidget):
    def __init__(self, fuzhi_jieguo, tray_icon):
        super().__init__()
        self.fuzhi_jieguo = fuzhi_jieguo
        self.tray_icon = tray_icon
        self.init()

    def init(self):
        self.setWindowTitle("负值解脱清单")
        self.resize(1200, 800)

        layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 上半部分：表格
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["投标人名称", "负值清单项个数"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # 填充表格
        bidders = list(self.fuzhi_jieguo.keys())
        self.table.setRowCount(len(bidders))
        for i, name in enumerate(bidders):
            count = 0
            for dx in self.fuzhi_jieguo[name].values():
                for dw in dx.values():
                    count += len(dw.get('分部清单', [])) + len(dw.get('措施清单', []))
            self.table.setItem(i, 0, QTableWidgetItem(name))
            self.table.setItem(i, 1, QTableWidgetItem(str(count)))

        # 下半部分：树状结构
        tree = QTreeWidget()
        tree.setHeaderLabels(["编码", "名称", "项目特征", "单位", "数量", "综合单价"])
        tree.setEditTriggers(QTreeWidget.EditTrigger.NoEditTriggers)
        tree.setColumnWidth(0, 150)
        tree.setColumnWidth(1, 200)
        tree.setColumnWidth(2, 400)
        tree.setColumnWidth(3, 60)
        tree.setColumnWidth(4, 80)
        tree.setColumnWidth(5, 100)
        tree.header().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        tree.setWordWrap(True)
        tree.setStyleSheet("QTreeWidget::item { height: 36px; }")

        # 点击表格切换树内容
        def on_table_clicked(row):
            tree.clear()
            name = bidders[row]
            root_data = self.fuzhi_jieguo[name]
            for dx_name, dx_data in root_data.items():
                dx_item = QTreeWidgetItem(tree, [dx_name, "", "", "", "", ""])
                dx_item.setForeground(0, Qt.GlobalColor.darkBlue)
                for dw_name, dw_data in dx_data.items():
                    dw_item = QTreeWidgetItem(dx_item, [dw_name, "", "", "", "", ""])
                    dw_item.setForeground(0, Qt.GlobalColor.darkGreen)

                    for qd in dw_data.get('分部清单', []):
                        item = QTreeWidgetItem(dw_item)
                        item.setData(0, Qt.ItemDataRole.DisplayRole, qd.get("编码", ""))
                        item.setData(1, Qt.ItemDataRole.DisplayRole, qd.get("名称", ""))
                        item.setData(2, Qt.ItemDataRole.DisplayRole, qd.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setData(3, Qt.ItemDataRole.DisplayRole, qd.get("单位", ""))
                        item.setData(4, Qt.ItemDataRole.DisplayRole, qd.get("数量", ""))
                        item.setData(5, Qt.ItemDataRole.DisplayRole, qd.get("综合单价", ""))

                    for cs in dw_data.get('措施清单', []):
                        item = QTreeWidgetItem(dw_item)
                        item.setData(0, Qt.ItemDataRole.DisplayRole, cs.get("编码", ""))
                        item.setData(1, Qt.ItemDataRole.DisplayRole, cs.get("名称", ""))
                        item.setData(2, Qt.ItemDataRole.DisplayRole, cs.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setData(3, Qt.ItemDataRole.DisplayRole, cs.get("单位", ""))
                        item.setData(4, Qt.ItemDataRole.DisplayRole, cs.get("数量", ""))
                        item.setData(5, Qt.ItemDataRole.DisplayRole, cs.get("综合单价", ""))

            tree.expandAll()

        if bidders:
            self.table.selectRow(0)
            on_table_clicked(0)
        self.table.clicked.connect(lambda idx: on_table_clicked(idx.row()))

        splitter.addWidget(self.table)
        splitter.addWidget(tree)
        splitter.setSizes([150, 650])
        layout.addWidget(splitter)

        # 新增导出按钮
        btn_export = QPushButton("导出有问题投标人清单")
        btn_export.clicked.connect(self.export_problem_bidders)
        btn_export.setDefault(True)
        layout.addWidget(btn_export)

    def export_problem_bidders(self):
        # 只导出有负值/零值的投标人
        problem_bidders = []
        for i in range(self.table.rowCount()):
            name_item = self.table.item(i, 0)
            count_item = self.table.item(i, 1)
            if not name_item or not count_item:
                continue
            name = name_item.text()
            count = int(count_item.text())
            if count > 0:
                problem_bidders.append(name)

        if not problem_bidders:
            QMessageBox.information(self, "提示", "所有投标人均无问题，无需导出。")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存问题清单",
            "单价为负值汇总表.xlsx" if "负值" in self.windowTitle() else "单价为0汇总表.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        for name in problem_bidders:
            ws = wb.create_sheet(title=name[:31])  # Excel sheet名限制31字符

            # 表头
            headers = ["单项工程", "单位工程", "清单类型", "编码", "名称", "项目特征", "单位", "数量", "综合单价"]
            ws.append(headers)

            root_data = self.fuzhi_jieguo[name]

            for dx_name, dx_data in root_data.items():
                for dw_name, dw_data in dx_data.items():
                    has_data = False
                    # 检查是否有分部或措施清单
                    if dw_data.get('分部清单') or dw_data.get('措施清单'):
                        has_data = True

                    if not has_data:
                        continue  # 跳过完全没问题的单位工程

                    # 导出分部清单
                    for qd in dw_data.get('分部清单', []):
                        row = [
                            dx_name,
                            dw_name,
                            "分部清单",
                            functions.clean_for_excel(qd.get("编码", "")),
                            functions.clean_for_excel(qd.get("名称", "")),
                            functions.clean_for_excel(qd.get("项目特征", "")),
                            functions.clean_for_excel(qd.get("单位", "")),
                            functions.clean_for_excel(qd.get("数量", "")),
                            functions.clean_for_excel(qd.get("综合单价", ""))
                        ]
                        ws.append(row)

                    # 导出措施清单
                    for cs in dw_data.get('措施清单', []):
                        row = [
                            dx_name,
                            dw_name,
                            "措施清单",
                            functions.clean_for_excel(cs.get("编码", "")),
                            functions.clean_for_excel(cs.get("名称", "")),
                            functions.clean_for_excel(cs.get("项目特征", "")),
                            functions.clean_for_excel(cs.get("单位", "")),
                            functions.clean_for_excel(cs.get("数量", "")),
                            functions.clean_for_excel(cs.get("综合单价", ""))
                        ]
                        ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.number_format = '@'  # 强制文本格式

        try:
            wb.save(file_path)
            self.tray_icon.showMessage(
                "导出成功",
                f"已导出 {len(problem_bidders)} 个投标人的问题清单（含树状层级）\n{file_path}",
                QSystemTrayIcon.MessageIcon.Information
            )
        except Exception as e:
            self.tray_icon.showMessage(
                "导出失败",
                f"保存文件时出错：\n{e}",
                QSystemTrayIcon.MessageIcon.Critical
            )


class DiffDelegate(QStyledItemDelegate):
    def __init__(self, diff_flags, parent=None):
        super().__init__(parent)
        self.diff_flags = diff_flags  # {(row,col): [True/False,...]}

    def paint(self, painter, option, index):
        text = index.data(Qt.ItemDataRole.DisplayRole) or ""
        painter.save()
        painter.setClipRect(option.rect)

        flags = self.diff_flags.get((index.row(), index.column()), [False] * len(text))
        fm = option.fontMetrics
        line_height = fm.height()
        y = option.rect.y() + 2

        x = option.rect.x()
        for line in text.split("\n"):
            line_x = x
            for i, ch in enumerate(line):
                w = fm.horizontalAdvance(ch)
                rect = option.rect
                rect.setX(line_x)
                rect.setY(y)
                rect.setWidth(w)
                rect.setHeight(line_height)

                if i < len(flags) and flags[i]:
                    painter.fillRect(rect, QColor(255, 255, 0))  # 黄色背景
                    painter.setPen(QColor(200, 0, 0))  # 红字
                else:
                    painter.setPen(option.palette.text().color())

                painter.drawText(rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, ch)
                line_x += w
            y += line_height
        painter.restore()


def compute_diffs(text1, text2):
    s = difflib.SequenceMatcher(None, text1, text2)
    diff1 = [False] * len(text1)
    diff2 = [False] * len(text2)
    for tag, i1, i2, j1, j2 in s.get_opcodes():
        if tag in ("replace", "delete"):
            for i in range(i1, i2):
                diff1[i] = True
        if tag in ("replace", "insert"):
            for j in range(j1, j2):
                diff2[j] = True
    return diff1, diff2


class qingdan_window(QWidget):
    def __init__(self, qingdan_jieguo, tray_icon):
        super().__init__()
        self.qingdan_jieguo = qingdan_jieguo
        self.tray_icon = tray_icon
        self.tree = None
        self.diff_flags = {}
        self.init()

    def init(self):
        self.setWindowTitle("清单差异对比")
        self.resize(1500, 900)

        layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Orientation.Vertical)

        # ───── 上：统计表 ─────
        table_stat = QTableWidget()
        table_stat.setColumnCount(2)
        table_stat.setHorizontalHeaderLabels(["投标人名称", "差异项个数"])
        table_stat.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table_stat.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        bidders = list(self.qingdan_jieguo.keys())
        table_stat.setRowCount(len(bidders))
        for i, name in enumerate(bidders):
            cnt = 0
            for dx in self.qingdan_jieguo[name].values():
                for dw in dx.values():
                    if dw.get('分部清单'): cnt += 1
                    if dw.get('措施清单'): cnt += 1
            table_stat.setItem(i, 0, QTableWidgetItem(name))
            table_stat.setItem(i, 1, QTableWidgetItem(str(cnt)))

        # ───── 下：差异表格（树形）─────
        self.tree = QTreeWidget()
        self.tree.setColumnCount(11)
        headers = [
            "层级",
            "招标-编码", "招标-名称", "招标-项目特征", "招标-单位", "招标-工程量",
            "投标-编码", "投标-名称", "投标-项目特征", "投标-单位", "投标-工程量"
        ]
        self.tree.setHeaderLabels(headers)
        self.tree.setEditTriggers(QTreeWidget.EditTrigger.NoEditTriggers)
        self.tree.setWordWrap(True)
        self.tree.setUniformRowHeights(False)  # 允许每行不同高度
        self.tree.setStyleSheet("QTreeWidget::item { height: 46px; }")
        self.tree.setColumnWidth(0, 200)
        for i in range(1, 11):
            self.tree.setColumnWidth(i, 150)

        # 表头颜色 + 加粗
        green = QBrush(QColor(0, 180, 0))
        orange = QBrush(QColor(255, 140, 0))
        bold = QFont()
        bold.setBold(True)
        header_item = QTreeWidgetItem(headers)
        self.tree.setHeaderItem(header_item)
        for col in range(1, 6):
            header_item.setForeground(col, green)
            header_item.setFont(col, bold)
        for col in range(6, 11):
            header_item.setForeground(col, orange)
            header_item.setFont(col, bold)

        # 差异委托
        self.delegate = DiffDelegate(self.diff_flags, self.tree)

        def fill_tree(row_idx):
            self.tree.clear()
            self.diff_flags.clear()
            if row_idx >= len(bidders):
                return
            name = bidders[row_idx]
            root = self.qingdan_jieguo[name]

            item_idx = 0
            for dx_name, dx_data in root.items():
                dx_item = QTreeWidgetItem(self.tree, [dx_name] + [""] * 10)
                dx_item.setBackground(0, QColor(240, 240, 240))
                for dw_name, dw_data in dx_data.items():
                    dw_item = QTreeWidgetItem(dx_item, [dw_name] + [""] * 10)

                    # 分部清单
                    fb = dw_data.get('分部清单')
                    if isinstance(fb, dict) and fb:
                        zb, tb = fb['招标清单'], fb['投标清单']
                        item = QTreeWidgetItem(dw_item, ["分部清单差异"] + [""] * 10)
                        item.setText(1, zb.get("编码", ""))
                        item.setText(2, zb.get("名称", ""))
                        item.setText(3, zb.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setText(4, zb.get("单位", ""))
                        item.setText(5, zb.get("数量", ""))
                        item.setText(6, tb.get("编码", ""))
                        item.setText(7, tb.get("名称", ""))
                        item.setText(8, tb.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setText(9, tb.get("单位", ""))
                        item.setText(10, tb.get("数量", ""))

                        for col_zb, col_tb, key in [(2, 7, "名称"), (3, 8, "项目特征"), (4, 9, "单位"),
                                                    (5, 10, "数量")]:
                            t1 = zb.get(key, "")
                            t2 = tb.get(key, "")
                            d1, d2 = compute_diffs(t1, t2)
                            self.diff_flags[(item_idx, col_zb)] = d1
                            self.diff_flags[(item_idx, col_tb)] = d2
                        item_idx += 1

                    # 措施清单
                    cs = dw_data.get('措施清单')
                    if isinstance(cs, dict) and cs:
                        zb, tb = cs['招标清单'], cs['投标清单']
                        item = QTreeWidgetItem(dw_item, ["措施清单差异"] + [""] * 10)
                        item.setText(1, zb.get("编码", ""))
                        item.setText(2, zb.get("名称", ""))
                        item.setText(3, zb.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setText(4, zb.get("单位", ""))
                        item.setText(5, zb.get("数量", ""))
                        item.setText(6, tb.get("编码", ""))
                        item.setText(7, tb.get("名称", ""))
                        item.setText(8, tb.get("项目特征", "").replace("\\r\\n", "\n"))
                        item.setText(9, tb.get("单位", ""))
                        item.setText(10, tb.get("数量", ""))

                        for col_zb, col_tb, key in [(2, 7, "名称"), (3, 8, "项目特征"), (4, 9, "单位"),
                                                    (5, 10, "数量")]:
                            t1 = zb.get(key, "")
                            t2 = tb.get(key, "")
                            d1, d2 = compute_diffs(t1, t2)
                            self.diff_flags[(item_idx, col_zb)] = d1
                            self.diff_flags[(item_idx, col_tb)] = d2
                        item_idx += 1

            self.tree.setItemDelegate(self.delegate)
            self.tree.expandAll()
            # 自动调整所有行高
            for i in range(self.tree.topLevelItemCount()):
                self.resize_row_recursive(self.tree.topLevelItem(i))

        def on_click(idx):
            fill_tree(idx.row())

        if table_stat.rowCount() > 0:
            table_stat.selectRow(0)
            fill_tree(0)
        else:
            self.tree.clear()
        table_stat.clicked.connect(on_click)

        splitter.addWidget(table_stat)
        splitter.addWidget(self.tree)
        splitter.setSizes([120, 780])
        layout.addWidget(splitter)

        # 新增导出按钮
        btn_export = QPushButton("导出清单不一致项")
        # btn_export.setStyleSheet("QPushButton { background: #FF5722; color: white; padding: 8px; font-weight: bold; }")
        btn_export.clicked.connect(self.export_diff_sheets)
        btn_export.setDefault(True)
        layout.addWidget(btn_export)

    def resize_row_recursive(self, item):
        row = self.tree.indexFromItem(item).row()
        if row >= 0:
            self.tree.resizeColumnToContents(row)
        for i in range(item.childCount()):
            self.resize_row_recursive(item.child(i))

    def export_diff_sheets(self):
        problem_bidders = [name for name in self.qingdan_jieguo.keys()
                           if any(dw.get('分部清单') or dw.get('措施清单')
                                  for dx in self.qingdan_jieguo[name].values()
                                  for dw in dx.values())]

        if not problem_bidders:
            QMessageBox.information(self, "提示", "所有投标人清单一致，无需导出。")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存清单差异", "清单不一致汇总表.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        wb.remove(wb.active)

        for name in problem_bidders:
            ws = wb.create_sheet(title=name[:31])
            headers = [
                "层级", "招标-编码", "招标-名称", "招标-项目特征", "招标-单位", "招标-工程量",
                "投标-编码", "投标-名称", "投标-项目特征", "投标-单位", "投标-工程量"
            ]
            ws.append(headers)

            root = self.qingdan_jieguo[name]
            for dx_name, dx_data in root.items():
                for dw_name, dw_data in dx_data.items():
                    for typ in ['分部清单', '措施清单']:
                        item = dw_data.get(typ)
                        if isinstance(item, dict) and item:
                            zb, tb = item['招标清单'], item['投标清单']
                            row = [
                                f"{dx_name} → {dw_name} → {typ}差异",
                                zb.get("编码", ""), zb.get("名称", ""), zb.get("项目特征", ""), zb.get("单位", ""),
                                zb.get("数量", ""),
                                tb.get("编码", ""), tb.get("名称", ""), tb.get("项目特征", ""), tb.get("单位", ""),
                                tb.get("数量", "")
                            ]
                            ws.append(row)

        try:
            wb.save(file_path)
            self.tray_icon.showMessage("成功", f"已导出 {len(problem_bidders)} 个投标人差异：\n{file_path}")
        except Exception as e:
            self.tray_icon.showMessage("错误", f"导出失败：\n{e}")


class qingDanCompareWindow(QDialog):
    def __init__(self, result: dict, tray_icon):
        super().__init__()
        self.result = result
        # 预处理数据：按清单类别分类
        self.categorized_data = self._categorize_data()
        self.setWindowTitle("投标清单差异对比工具")
        self.resize(1700, 1000)
        self.tray_icon = tray_icon
        self.init_ui()
        self.get_company_shuliang_data()
        self.update_uptable_data()

    def _categorize_data(self):
        """将数据按清单类别分类（动态生成所有类别）"""
        categorized = {}
        for company, items in self.result.items():
            categorized[company] = {}  # 改成空 dict，不预设 key
            for item in items:
                category = item.get('清单类别', '其他项目')
                if category not in categorized[company]:
                    categorized[company][category] = []  # 动态创建
                categorized[company][category].append(item)
        return categorized

    def get_company_shuliang_data(self):
        self.company_shuliang_data = []
        for company_name, value in self.result.items():
            a_list = [company_name, len(value)]
            self.company_shuliang_data.append(a_list)

        self.update_uptable_data()

    def update_uptable_data(self):
        for row, item in enumerate(self.company_shuliang_data):
            for col, value in enumerate(item):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.up_table.setItem(row, col, item)

    def update_mingxi_table(self, company_name, kind):
        self.kind_data = self.categorized_data.get(company_name, {}).get(kind, [])
        if not self.kind_data:
            self.down_left_table.setRowCount(0)
            self.down_right_table.setRowCount(0)
            return

        sample_tender = self.kind_data[0].get("招标清单", {})
        sample_bid = self.kind_data[0].get("投标清单", {})
        tender_keys = list(sample_tender.keys())
        bid_keys = list(sample_bid.keys())

        left_headers = ["单项工程", "单位工程"] + tender_keys
        right_headers = bid_keys

        self.down_left_table.clear()
        self.down_left_table.setColumnCount(len(left_headers))
        self.down_left_table.setHorizontalHeaderLabels(left_headers)
        self.down_left_table.setRowCount(len(self.kind_data))

        self.down_right_table.clear()
        self.down_right_table.setColumnCount(len(right_headers))
        self.down_right_table.setHorizontalHeaderLabels(right_headers)
        self.down_right_table.setRowCount(len(self.kind_data))

        # 设置HTML代理以支持富文本
        self.down_right_table.setItemDelegate(HtmlDelegate())

        for row, item in enumerate(self.kind_data):
            # 填充左侧表格（招标清单）
            self.down_left_table.setItem(row, 0, QTableWidgetItem(item.get("单项工程", "")))
            self.down_left_table.setItem(row, 1, QTableWidgetItem(item.get("单位工程", "")))
            for col, key in enumerate(tender_keys):
                val = str(item.get("招标清单", {}).get(key, ""))
                self.down_left_table.setItem(row, col + 2, QTableWidgetItem(val))

            # 填充右侧表格（投标清单）并标记差异
            for col, key in enumerate(bid_keys):
                tender_val = str(item.get("招标清单", {}).get(key, ""))
                bid_val = str(item.get("投标清单", {}).get(key, ""))
                marked_text = bid_val
                if key in sample_tender:
                    matcher = difflib.SequenceMatcher(None, tender_val, bid_val)
                    marked_parts = []
                    for opcode, a1, a2, b1, b2 in matcher.get_opcodes():
                        if opcode == 'equal':
                            marked_parts.append(bid_val[b1:b2])
                        elif opcode == 'insert':
                            marked_parts.append(f'<font color="red">{bid_val[b1:b2]}</font>')
                        elif opcode == 'delete':
                            marked_parts.append(f'<font color="red">[{tender_val[a1:a2]}]</font>')
                        elif opcode == 'replace':
                            marked_parts.append(f'<font color="red">{bid_val[b1:b2]}</font>')
                    marked_text = ''.join(marked_parts)
                self.down_right_table.setItem(row, col, QTableWidgetItem(marked_text))

    def init_ui(self):
        main_layout = QVBoxLayout()
        main_splitter = QSplitter(Qt.Orientation.Vertical)

        # 上部分：公司列表表格
        self.up_table = QTableWidget()
        headers = ["公司名称", "错误数量"]
        self.up_table.setColumnCount(len(headers))
        self.up_table.setHorizontalHeaderLabels(headers)
        self.up_table.setRowCount(len(self.result))

        # 让表格占满整个上部区域
        # self.up_table.setSizeAdjustPolicy(QTableWidget.Ad)  # 可选，配合布局更稳
        self.up_table.horizontalHeader().setStretchLastSection(True)  # 最后一列拉伸填满

        # 关键：让所有列等宽且占满宽度（推荐方式）
        self.up_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.up_table.itemClicked.connect(self.on_company_selected)

        # 下部分：左右布局
        down_widget = QWidget()
        down_layout = QHBoxLayout()
        down_layout.setContentsMargins(0, 0, 0, 0)
        down_layout.setSpacing(5)

        # 左侧：清单类别树形视图
        self.down_tree_view = QTreeView()
        self.down_tree_view.setMinimumWidth(200)
        self.down_tree_view.setMaximumWidth(300)

        # 创建树模型
        self.tree_model = QStandardItemModel()
        self.tree_model.setHorizontalHeaderLabels(["清单类别"])
        self.down_tree_view.setModel(self.tree_model)
        self.down_tree_view.clicked.connect(self.on_category_selected)

        # 右侧：分割器包含两个表格
        down_splitter = QSplitter(Qt.Orientation.Horizontal)

        # 控制价表格
        self.down_left_table = QTableWidget()

        # 投标数据表格
        self.down_right_table = QTableWidget()

        down_splitter.addWidget(self.down_left_table)
        down_splitter.addWidget(self.down_right_table)
        down_splitter.setStretchFactor(0, 6)
        down_splitter.setStretchFactor(1, 4)

        # 添加到布局
        down_layout.addWidget(self.down_tree_view, 2)  # 20%
        down_layout.addWidget(down_splitter, 7)  # 80%
        down_widget.setLayout(down_layout)

        # 下载按钮
        self.export_excel_button = QPushButton("导出")
        self.export_excel_button.setDefault(True)
        self.export_excel_button.clicked.connect(self.export_excel)

        # 主分割器
        main_splitter.addWidget(self.up_table)
        main_splitter.addWidget(down_widget)
        main_splitter.addWidget(self.export_excel_button)
        main_splitter.setStretchFactor(0, 1)
        main_splitter.setStretchFactor(1, 2)
        main_splitter.setStretchFactor(2, 1)

        main_layout.addWidget(main_splitter)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setLayout(main_layout)

        # 初始化默认选择
        if self.result:
            first_company = list(self.result.keys())[0]
            self.current_company = first_company
            self.populate_category_tree(first_company)

            # 自动选择第一个类别
            if self.tree_model.rowCount() > 0:
                first_index = self.tree_model.index(0, 0)
                # 获取类别名称
                first_item = self.tree_model.itemFromIndex(first_index)
                category_name = first_item.text()
                self.update_mingxi_table(self.current_company, category_name)

    def populate_category_tree(self, company_name):
        """填充清单类别树"""
        self.tree_model.clear()
        self.tree_model.setHorizontalHeaderLabels(["清单类别"])

        root = self.tree_model.invisibleRootItem()  # 获取根项

        self.kind_list = []
        for name, value in self.result.items():
            for item in value:
                if name == company_name:
                    self.kind_list.append(item.get('清单类别'))

        self.kind_list = set(self.kind_list)

        for item in self.kind_list:
            standard_item = QStandardItem(item)
            root.appendRow(standard_item)

        # 展开所有节点
        self.down_tree_view.expandAll()

    def on_company_selected(self, item):
        """当选择公司时触发"""
        if not item:
            return

        company_name = self.up_table.item(item.row(), 0).text()
        self.current_company = company_name

        # 更新类别树
        self.populate_category_tree(company_name)

        # 自动选择第一个类别并加载数据
        if self.tree_model.rowCount() > 0:
            first_index = self.tree_model.index(0, 0)
            self.down_tree_view.setCurrentIndex(first_index)

            # 获取类别名称
            first_item = self.tree_model.itemFromIndex(first_index)
            category_name = first_item.text()
            self.load_category_data(company_name, category_name)

    def on_category_selected(self, index):
        """当选择清单类别时触发"""
        if not index.isValid():
            return

        item = self.tree_model.itemFromIndex(index)
        if item:
            self.load_category_data(self.current_company, item.text())


    def load_category_data(self, company_name, category_name):
        """加载指定公司和类别的数据"""
        if company_name not in self.categorized_data:
            return

        self.update_mingxi_table(company_name=company_name, kind=category_name)


        # 优化表格显示
        self.optimize_table_display()

    def optimize_table_display(self):
        """优化表格显示效果"""
        # 设置列宽
        self.down_left_table.setColumnWidth(4, 300)  # 项目特征列
        self.down_right_table.setColumnWidth(2, 300)  # 项目特征列

        # 设置自适应行高
        self.down_left_table.resizeRowsToContents()
        self.down_right_table.resizeRowsToContents()

        # 设置垂直表头自适应内容高度
        self.down_left_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.down_right_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)

        # 设置交替行颜色
        self.down_left_table.setAlternatingRowColors(True)
        self.down_right_table.setAlternatingRowColors(True)

        # 设置自动换行
        self.down_left_table.setWordWrap(True)
        self.down_right_table.setWordWrap(True)

        # 设置水平表头自适应
        self.down_left_table.horizontalHeader().setStretchLastSection(True)
        self.down_right_table.horizontalHeader().setStretchLastSection(True)

    def export_excel(self):
        # 弹出文件夹选择对话框，选择保存位置
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "选择保存文件夹",
            "",  # 默认路径，空字符串表示当前目录
        )

        if not folder_path:
            return  # 用户取消保存

        # 遍历每个投标人（公司）
        for company_name, company_data in self.categorized_data.items():
            # 创建新的工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 移除默认的工作表

            # 遍历该投标人的每个类别
            for category_name, items in company_data.items():
                if not items:  # 如果类别为空则跳过
                    continue

                # 创建新的工作表，使用类别名称作为表名
                ws = wb.create_sheet(title=category_name)

                # 处理每个类别中的数据
                rows_data = []

                for item in items:
                    row_data = {}

                    # 添加公共字段
                    if '单项工程' in item:
                        row_data['单项工程'] = item['单项工程']
                    if '单位工程' in item:
                        row_data['单位工程'] = item['单位工程']
                    if '清单类别' in item:
                        row_data['清单类别'] = item['清单类别']

                    # 处理招标清单和投标清单
                    for list_type in ['招标清单', '投标清单']:
                        if list_type in item:
                            list_data = item[list_type]
                            for key, value in list_data.items():
                                # 为招标和投标数据添加前缀以便区分
                                prefix = "招标_" if list_type == "招标清单" else "投标_"
                                row_data[prefix + key] = value

                    rows_data.append(row_data)

                # 将数据转换为DataFrame
                if rows_data:
                    df = pd.DataFrame(rows_data)

                    # 添加表头
                    ws.append(list(df.columns))

                    # 写入数据行
                    for r in dataframe_to_rows(df, index=False, header=False):
                        ws.append(r)

                    # 自动调整列宽
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass

                        adjusted_width = min(max_length + 2, 50)  # 设置最大宽度为50
                        ws.column_dimensions[column_letter].width = adjusted_width

                    # 冻结首行（标题行）
                    ws.freeze_panes = ws['A2']

            # 生成文件路径：文件夹路径/公司名称.xlsx
            # 确保文件名合法（移除特殊字符）
            safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '_', '-')).strip()
            safe_company_name = safe_company_name.replace(' ', '_')  # 将空格替换为下划线

            # 如果安全名称为空，使用默认名称
            if not safe_company_name:
                safe_company_name = f"公司_{len(os.listdir(folder_path)) + 1}"

            file_path = os.path.join(folder_path, f"{safe_company_name}.xlsx")

            # 如果文件已存在，添加序号
            counter = 1
            original_path = file_path
            while os.path.exists(file_path):
                file_path = original_path.replace('.xlsx', f'_{counter}.xlsx')
                counter += 1

            # 保存工作簿
            try:
                wb.save(file_path)
                print(f"已保存: {file_path}")

            except Exception as e:
                print(f"保存文件时出错: {e}")
        self.tray_icon.showMessage("保存成功", f"已导出清单不一致结果。")



class LockCheckWindow(QWidget):
    def __init__(self, table_data, tray_icon):
        super().__init__()
        self.table_data = table_data
        self.tray_icon = tray_icon
        self.init_ui()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        headers = [
            "序号", "投标人", "总价",
            "加密锁数量", "锁号", "重复单位(锁)",
            "MAC数量", "物理地址", "重复单位(MAC)"
        ]
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # 导出按钮
        btn_layout = QHBoxLayout()
        btn_export = QPushButton("导出为 Excel")

        btn_export.clicked.connect(self.export_excel)
        # btn_layout.addStretch()
        btn_layout.addWidget(btn_export)

        layout.addWidget(self.table)
        layout.addLayout(btn_layout)

        # 统计重复数量并显示在标题
        dup_count = sum(1 for r in self.table_data if "无" not in r["重复单位(锁)"] or "无" not in r["重复单位(MAC)"])
        status = f"发现 {dup_count} 处重复" if dup_count else "全部正常，无重复"
        self.setWindowTitle(f"加密锁检查 - {status}")

    def load_data(self):
        self.table.setRowCount(len(self.table_data))
        for row_idx, row in enumerate(self.table_data):
            for col_idx, key in enumerate([
                "序号", "投标人", "总价",
                "加密锁数量", "锁号", "重复单位(锁)",
                "MAC数量", "物理地址", "重复单位(MAC)"
            ]):
                item = QTableWidgetItem(row[key])
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)

                # 高亮重复项（红色加粗）
                if key in ("重复单位(锁)", "重复单位(MAC)") and "无" not in row[key]:
                    item.setForeground(Qt.GlobalColor.red)
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)

                self.table.setItem(row_idx, col_idx, item)

    def export_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "加密锁复核结果"

        headers = [
            "序号", "投标人", "总价",
            "加密锁数量", "锁号", "与此单位重复加密锁号的单位",
            "MAC地址数量", "物理地址", "与此单位重复物理地址的单位"
        ]
        ws.append(headers)

        for row in self.table_data:
            ws.append([
                row["序号"], row["投标人"], row["总价"],
                row["加密锁数量"], row["锁号"], row["重复单位(锁)"],
                row["MAC数量"], row["物理地址"], row["重复单位(MAC)"]
            ])

        file_path, _ = QFileDialog.getSaveFileName(self, "保存 Excel 文件", "硬件检测结果", "Excel 文件 (*.xlsx)")

        if file_path:
            try:
                wb.save(file_path)
                self.tray_icon.showMessage("成功", f"已导出表格：\n{file_path}")
            except Exception as e:
                self.tray_icon.showMessage("错误", f"导出失败：\n{e}")


class PianchaWindow(QWidget):
    def __init__(self, piancha_jieguo, pianchalv, tray_icon):
        super().__init__()
        self.data = piancha_jieguo
        self.tray_icon = tray_icon
        self.pianchalv = pianchalv
        self.init_ui()
        self.fill_all()

    def init_ui(self):
        self.setWindowTitle(f"±{self.pianchalv}% 偏差率清单检查")
        self.resize(1600, 900)

        layout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 上部：统计表（原生风格）
        self.stat_table = QTableWidget()
        self.stat_table.setColumnCount(2)
        self.stat_table.setHorizontalHeaderLabels(["投标人", "偏差数量"])
        self.stat_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.stat_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stat_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.stat_table.setAlternatingRowColors(True)

        # 下部：树形清单（极简原生）
        self.tree = QTreeWidget()
        headers = [
            "单项工程", "单位工程", "编码", "名称", "项目特征",
            "单位", "工程量", "控制价单价", "投标下浮率",
            "基准价", "投标单价", "偏差率", "备注"
        ]
        self.tree.setHeaderLabels(headers)
        self.tree.setColumnCount(len(headers))
        self.tree.setWordWrap(True)
        self.tree.setUniformRowHeights(False)
        self.tree.header().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

        # 关键：项目特征列固定宽度 300 + 允许换行
        self.tree.setColumnWidth(4, 300)  # 项目特征列
        for i in range(len(headers)):
            if i != 4:
                self.tree.header().setSectionResizeMode(i, QHeaderView.ResizeMode.Interactive)
        self.tree.header().setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)

        # 导出按钮（简洁红）
        btn_export = QPushButton("导出偏差清单")
        btn_export.clicked.connect(self.export_all)
        self.stat_table.clicked.connect(lambda idx: self.fill_tree(idx.row()))

        splitter.addWidget(self.stat_table)
        splitter.addWidget(self.tree)
        splitter.setSizes([140, 760])

        layout.addWidget(splitter)
        layout.addWidget(btn_export)

    def fill_all(self):
        bidders = list(self.data.keys())
        self.stat_table.setRowCount(len(bidders))
        self.stat_counts = {}

        for row, bidder in enumerate(bidders):
            total = 0
            for dx in self.data[bidder].values():
                for dw in dx.values():
                    total += len(dw.get(f'大于{self.pianchalv}%', [])) + len(dw.get(f'小于-{self.pianchalv}%', []))
            self.stat_counts[bidder] = total

            self.stat_table.setItem(row, 0, QTableWidgetItem(bidder))
            self.stat_table.setItem(row, 1, QTableWidgetItem(str(total)))

        if bidders:
            self.stat_table.selectRow(0)
            self.fill_tree(0)

        total_all = sum(self.stat_counts.values())
        self.setWindowTitle(f"±{self.pianchalv}% 偏差率清单 - 共 {total_all} 项")

    def fill_tree(self, row):
        bidder = self.stat_table.item(row, 0).text()
        if self.stat_counts.get(bidder, 0) == 0:
            self.tree.clear()
            return

        self.tree.clear()
        root = self.data[bidder]

        for dx_name, dx_data in root.items():
            dx_item = QTreeWidgetItem(self.tree, [dx_name])
            for dw_name, dw_data in dx_data.items():
                dw_item = QTreeWidgetItem(dx_item, ["", dw_name])

                items = (dw_data.get(f'大于{self.pianchalv}%', []) +
                         dw_data.get(f'小于-{self.pianchalv}%', []))

                for it in items:
                    remark = ""
                    if it['偏差率'] > self.pianchalv:
                        remark = f"偏高 {it['偏差率']:.1f}%"
                    elif it['偏差率'] < -self.pianchalv:
                        remark = f"偏低 {abs(it['偏差率']):.1f}%"

                    child = QTreeWidgetItem(dw_item, [
                        "", "",
                        it["编码"],
                        it["名称"],
                        it["项目特征"].replace("\\r\\n", "\n"),
                        it["单位"],
                        it["工程量"],
                        f"{it['控制价单价']:.2f}",
                        it["下浮率"],
                        f"{it['基准价']:.2f}",
                        f"{it['投标单价']:.2f}",
                        f"{it['偏差率']:+.2f}%",
                        remark
                    ])

        self.tree.expandAll()
        for i in range(self.tree.columnCount()):
            if i != 4:
                self.tree.resizeColumnToContents(i)

    def export_all(self):
        if not any(self.stat_counts.values()):
            QMessageBox.information(self, "提示", "无偏差项")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出偏差清单",
            f"±{self.pianchalv}%偏差率清单.xlsx",
            "Excel 文件 (*.xlsx)"
        )
        if not file_path: return

        wb = Workbook()
        wb.remove(wb.active)

        def clean(t): return str(t).replace('_x000D_', '\n').strip()

        for bidder, total in self.stat_counts.items():
            if total == 0: continue
            ws = wb.create_sheet(title=bidder[:31])
            # 添加一个标题行，内容是公司名称
            ws.append([bidder])
            headers = [
                "单项工程", "单位工程", "编码", "名称", "项目特征",
                "单位", "工程量", "控制价单价", "投标下浮率",
                "基准价", "投标单价", "偏差率", "备注"
            ]
            ws.append(headers)

            for dx_name, dx_data in self.data[bidder].items():
                for dw_name, dw_data in dx_data.items():
                    items = dw_data.get(f'大于{self.pianchalv}%', []) + dw_data.get(f'小于-{self.pianchalv}%', [])
                    for it in items:
                        remark = "偏高" if it['偏差率'] > self.pianchalv else "偏低"
                        remark += f" {abs(it['偏差率']):.1f}%"
                        ws.append([
                            dx_name, dw_name,
                            it["编码"], it["名称"], clean(it["项目特征"]),
                            it["单位"], it["工程量"],
                            it["控制价单价"], it["下浮率"],
                            it["基准价"], it["投标单价"],
                            f"{it['偏差率']:+.2f}%", remark
                        ])

            for cell in ws["E:E"]: cell.alignment = Alignment(wrap_text=True, vertical="top")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.col_idx >= 3: cell.number_format = '@'

        wb.save(file_path)
        self.tray_icon.showMessage("成功", f"已导出 {file_path}")


class zbfx_window(QDialog):
    def __init__(self, zb_data: dict, tb_data: list, kzj_data: dict):
        super().__init__()
        self.setWindowTitle("中标分析")
        self.setGeometry(100, 100, 800, 600)
        self.setWindowIcon(QIcon(resource_path("icons/zbfx.png")))
        self.tb_data = tb_data
        self.zb_data = zb_data
        self.kzj_data = kzj_data
        # 初始化投标单位数据
        self.names = []
        for tb in self.tb_data:
            # 跳过有效性不为真的
            if not tb['投标信息'].get("有效性", True):
                continue
            self.names.append(tb['投标信息']['投标人'])

        self.dx_data = {}
        self.piangao_data = {}
        self.piandi_data = {}
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        bili_layout = QHBoxLayout()
        kzj_label = QLabel("控制价比例:")
        self.kzj_bili_input = QLineEdit()
        self.kzj_bili_input.setValidator(QDoubleValidator(0, 100, 2, self))
        self.kzj_bili_input.setText("40")
        self.kzj_bili_input_unit = QLabel("%")
        bili_layout.addWidget(kzj_label)
        bili_layout.addWidget(self.kzj_bili_input)
        bili_layout.addWidget(self.kzj_bili_input_unit)

        select_zbdw_layout = QHBoxLayout()
        select_zbdw_label = QLabel("选择中标单位:")
        self.zbdw_select = QComboBox()
        self.zbdw_select.addItems(self.names)
        select_zbdw_layout.addWidget(select_zbdw_label)
        select_zbdw_layout.addWidget(self.zbdw_select)

        # 设置偏差率
        pianchalv_layout = QHBoxLayout()
        pianchalv_label = QLabel("偏差率:")
        self.pianchalv_input = QLineEdit()
        self.pianchalv_input.setValidator(QDoubleValidator(0, 100, 2, self))
        self.pianchalv_input.setText("10")
        danwei_label_1 = QLabel("%")
        self.condition_combo = QComboBox()
        self.condition_combo.addItems(["或", "且"])
        self.condition_combo.setCurrentText("且")
        pianchazhi_label = QLabel("偏差值:")
        self.pianchazhi_input = QLineEdit()
        self.pianchazhi_input.setValidator(QDoubleValidator(0, 100, 2, self))
        self.pianchazhi_input.setText("100000")
        danwei_label_2 = QLabel("元")

        # 设置一个勾选项，单价中是否包含设备费
        # 创建勾选框组件
        self.include_shebei_checkbox = QCheckBox("单价含设备")
        self.include_shebei_checkbox.setChecked(False)

        pianchalv_layout.addWidget(pianchalv_label)
        pianchalv_layout.addWidget(self.pianchalv_input)
        pianchalv_layout.addWidget(danwei_label_1)
        pianchalv_layout.addWidget(self.condition_combo)
        pianchalv_layout.addWidget(pianchazhi_label)
        pianchalv_layout.addWidget(self.pianchazhi_input)
        pianchalv_layout.addWidget(danwei_label_2)
        pianchalv_layout.addWidget(self.include_shebei_checkbox)

        btns_layout = QHBoxLayout()
        self.start_btn = QPushButton("开始分析")
        self.start_btn.clicked.connect(self.start_analysis)

        self.generate_btn = QPushButton("生成附件表格")
        self.generate_btn.clicked.connect(self.generate_report)
        btns_layout.addWidget(self.start_btn)
        btns_layout.addWidget(self.generate_btn)

        self.zbfx_tabs = QTabWidget()

        self.danxiang_table = QTableWidget()
        self.danxiang_table_titles = ["项目名称", "拟中标价", "基准价", "价格偏差", "偏差率"]
        self.danxiang_table.setColumnCount(len(self.danxiang_table_titles))
        self.danxiang_table.setHorizontalHeaderLabels(self.danxiang_table_titles)
        self.zbfx_tabs.addTab(self.danxiang_table, "单项工程中标报价对比分析表")

        self.chaochu_table = QTableWidget()
        self.chaochu_table_titles = ["编码", "名称", "项目特征", "单位", "工程量", "控制价单价", "投标平均值", "基准价",
                                     "投标单价", "单价偏差值", "单价偏差比例", "合价偏差值"]
        self.chaochu_table.setColumnCount(len(self.chaochu_table_titles))
        self.chaochu_table.setHorizontalHeaderLabels(self.chaochu_table_titles)
        self.zbfx_tabs.addTab(self.chaochu_table, "偏高项")

        self.diyu_table = QTableWidget()
        self.diyu_table_titles = ["编码", "名称", "项目特征", "单位", "工程量", "控制价单价", "投标平均值", "基准价",
                                  "投标单价", "单价偏差值", "单价偏差比例", "合价偏差值"]
        self.diyu_table.setColumnCount(len(self.diyu_table_titles))
        self.diyu_table.setHorizontalHeaderLabels(self.diyu_table_titles)
        self.zbfx_tabs.addTab(self.diyu_table, "偏低项")

        layout.addLayout(bili_layout)
        layout.addLayout(select_zbdw_layout)
        layout.addLayout(pianchalv_layout)
        # layout.addLayout(pianchazhi_layout)
        layout.addLayout(btns_layout)
        layout.addWidget(self.zbfx_tabs)

        self.start_btn.setShortcut("Ctrl+B")
        self.generate_btn.setShortcut("Ctrl+R")

        self.setLayout(layout)

    def generate_report(self):
        if not self.dx_data and not self.list_items:
            QMessageBox.warning(self, "警告", "未进行分析，无数据可生成")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "保存报告", "中标报价分析报告.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        wb = Workbook()
        wb.remove(wb.active)

        # 字体和对齐方式
        title_font = Font(name='Arial', size=12, bold=True)
        content_font = Font(name='Arial', size=10)
        title_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        # A4纸张设置（以点为单位，1点=1/72英寸，A4宽度约595点，高度842点）
        def set_sheet_format(ws, headers):
            ws.print_options.horizontalCentered = True
            ws.print_options.verticalCentered = True
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_margins.left = ws.page_margins.right = 0.5  # 左右边距
            ws.page_margins.top = ws.page_margins.bottom = 0.75  # 上下边距

            # 设置列宽以适应A4纸张
            total_width = 595 - 0.5 * 2 * 72  # 减去左右边距（英寸转点）
            col_count = len(headers)
            col_width = total_width / col_count / 7.5  # 粗略估算每列宽度（点转字符宽度）
            for col in range(1, col_count + 1):
                ws.column_dimensions[get_column_letter(col)].width = col_width

            # 设置行高
            ws.row_dimensions[1].height = 30  # 标题行高
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 25  # 内容行高

        # 附表一
        ws1 = wb.create_sheet("附表一 单项工程中标报价对比分析表")
        ws1.append(self.danxiang_table_titles)
        for row in range(self.danxiang_table.rowCount()):
            row_data = [self.danxiang_table.item(row, col).text() if self.danxiang_table.item(row, col) else "" for col
                        in range(self.danxiang_table.columnCount())]
            ws1.append(row_data)

        set_sheet_format(ws1, self.danxiang_table_titles)
        for row in ws1.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = title_font
                cell.alignment = title_align
                cell.border = thin_border
        for row in ws1.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font
                cell.alignment = content_align
                cell.border = thin_border

        # 附表二
        ws2 = wb.create_sheet("附表二 基准价偏差 超高项")
        ws2.append(self.chaochu_table_titles)
        for row in range(self.chaochu_table.rowCount()):
            row_data = [self.chaochu_table.item(row, col).text() if self.chaochu_table.item(row, col) else "" for col in
                        range(self.chaochu_table.columnCount())]
            ws2.append(row_data)

        set_sheet_format(ws2, self.chaochu_table_titles)
        for row in ws2.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = title_font
                cell.alignment = title_align
                cell.border = thin_border
        for row in ws2.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font
                cell.alignment = content_align
                cell.border = thin_border

        # 附表三
        ws3 = wb.create_sheet("附表三 基准价偏差 超低项")
        ws3.append(self.diyu_table_titles)
        for row in range(self.diyu_table.rowCount()):
            row_data = [self.diyu_table.item(row, col).text() if self.diyu_table.item(row, col) else "" for col in
                        range(self.diyu_table.columnCount())]
            ws3.append(row_data)

        set_sheet_format(ws3, self.diyu_table_titles)
        for row in ws3.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = title_font
                cell.alignment = title_align
                cell.border = thin_border
        for row in ws3.iter_rows(min_row=2):
            for cell in row:
                cell.font = content_font
                cell.alignment = content_align
                cell.border = thin_border

        wb.save(file_path)
        QMessageBox.information(self, "提示", "报告生成成功")

    def start_analysis(self):
        temp_dict = {}
        for tb in self.tb_data:
            if not tb['投标信息'].get("有效性", True):
                continue
            for item in tb["单项工程"]:
                name = item['名称']
                price = float(item['金额'])
                if name not in temp_dict:
                    temp_dict[name] = []
                temp_dict[name].append(price)

        self.dx_data = {item: [mean(data)] for item, data in temp_dict.items()}

        self.is_incloude_shebei = self.include_shebei_checkbox.isChecked()
        if self.is_incloude_shebei:
            key = "综合单价_含设备"
        else:
            key = "综合单价"

        self.list_items = {}
        for tb in self.tb_data:
            if not tb['投标信息'].get("有效性", True):
                continue
            for sp in tb["单项工程"]:
                for ue in sp['单位工程']:
                    for li in ue['分部清单']:
                        code = li['编码']
                        if code not in self.list_items:
                            self.list_items[code] = {
                                '名称': li['名称'],
                                '项目特征': li['项目特征'],
                                '单位': li['单位'],
                                '数量': li['数量'],
                                '单价': [],
                                '控制价单价': 0
                            }
                        self.list_items[code]['单价'].append(float(li[key]))
                    for cuoshi in ue['措施清单']:
                        code = cuoshi['编码']
                        if code not in self.list_items:
                            self.list_items[code] = {
                                '名称': cuoshi['名称'],
                                '项目特征': cuoshi['项目特征'],
                                '单位': cuoshi['单位'],
                                '数量': cuoshi['数量'],
                                '单价': [],
                                '控制价单价': 0
                            }

        for sp in self.kzj_data.get("单项工程", []):
            for ue in sp['单位工程']:
                for li in ue['分部清单']:
                    code = li['编码']
                    if code in self.list_items:
                        self.list_items[code]['控制价单价'] = float(li[key])

        self.update_huizong_table()
        self.update_piangao_table(key)
        self.update_piandi_table(key)

    def update_huizong_table(self):
        kzj_ratio = float(self.kzj_bili_input.text()) / 100 if self.kzj_bili_input.text() else 0.4
        selected_zbdw = self.zbdw_select.currentText()
        table_data = []
        total_zbdw_price = total_base_price = total_price_diff = 0

        for item, data in self.dx_data.items():
            avg_price = data[0]
            zbdw_price = 0
            for tb in self.tb_data:
                if tb['投标信息'].get('投标人') == selected_zbdw:
                    for proj in tb["单项工程"]:
                        if proj['名称'] == item:
                            zbdw_price = float(proj['金额'])
                            break
                    break
            kzj_price = 0
            for kzj_item in self.kzj_data.get("单项工程", []):
                if kzj_item['名称'] == item:
                    kzj_price = float(kzj_item['金额'])
                    break
            base_price = kzj_price * kzj_ratio + avg_price * (1 - kzj_ratio)
            price_diff = zbdw_price - base_price
            deviation_rate = (price_diff / base_price * 100) if base_price != 0 else 0
            table_data.append(
                [item, f"{zbdw_price:.2f}", f"{base_price:.2f}", f"{price_diff:.2f}", f"{deviation_rate:.2f}%"])
            total_zbdw_price += zbdw_price
            total_base_price += base_price
            total_price_diff += price_diff

        total_deviation_rate = (total_price_diff / total_base_price * 100) if total_base_price != 0 else 0
        table_data.append(["汇总", f"{total_zbdw_price:.2f}", f"{total_base_price:.2f}", f"{total_price_diff:.2f}",
                           f"{total_deviation_rate:.2f}%"])

        self.danxiang_table.setRowCount(len(table_data))
        for row_idx, row_data in enumerate(table_data):
            for col_idx, value in enumerate(row_data):
                self.danxiang_table.setItem(row_idx, col_idx, QTableWidgetItem(value))

    def update_piangao_table(self, key):
        kzj_ratio = float(self.kzj_bili_input.text()) / 100 if self.kzj_bili_input.text() else 0.4
        selected_zbdw = self.zbdw_select.currentText()
        pianchalv = float(self.pianchalv_input.text()) if self.pianchalv_input.text() else 10
        pianchazhi = float(self.pianchazhi_input.text()) if self.pianchazhi_input.text() else 100000
        condition = self.condition_combo.currentText()

        table_data = []
        bid_unit_prices = {}
        for tb in self.tb_data:
            if tb['投标信息'].get('投标人') == selected_zbdw:
                for sp in tb["单项工程"]:
                    for ue in sp['单位工程']:
                        for li in ue['分部清单']:
                            bid_unit_prices[li['编码']] = float(li[key])
                break

        for code, info in self.list_items.items():
            avg_price = mean(info['单价']) if info['单价'] else 0
            kzj_price = info['控制价单价']
            base_price = kzj_price * kzj_ratio + avg_price * (1 - kzj_ratio)
            bid_price = bid_unit_prices.get(code, 0)
            unit_diff = bid_price - base_price
            unit_rate = (unit_diff / base_price * 100) if base_price != 0 else 0
            shuliang = float(info['数量'])
            total_diff = unit_diff * shuliang
            if unit_diff > 0 and (condition == "或" and (unit_rate > pianchalv or total_diff > pianchazhi) or
                                  condition == "且" and unit_rate > pianchalv and total_diff > pianchazhi):
                table_data.append([
                    code, info['名称'], info['项目特征'], info['单位'], f"{shuliang:.2f}",
                    f"{kzj_price:.2f}", f"{avg_price:.2f}", f"{base_price:.2f}",
                    f"{bid_price:.2f}", f"{unit_diff:.2f}", f"{unit_rate:.2f}%", f"{total_diff:.2f}"
                ])

        if not table_data:
            table_data.append(["无偏高项"] + [""] * (len(self.chaochu_table_titles) - 1))

        self.chaochu_table.setRowCount(len(table_data))
        for row_idx, row_data in enumerate(table_data):
            for col_idx, value in enumerate(row_data):
                self.chaochu_table.setItem(row_idx, col_idx, QTableWidgetItem(value))

    def update_piandi_table(self, key):
        kzj_ratio = float(self.kzj_bili_input.text()) / 100 if self.kzj_bili_input.text() else 0.4
        selected_zbdw = self.zbdw_select.currentText()
        pianchalv = float(self.pianchalv_input.text()) if self.pianchalv_input.text() else 10
        pianchazhi = float(self.pianchazhi_input.text()) if self.pianchazhi_input.text() else 100000
        condition = self.condition_combo.currentText()

        table_data = []
        bid_unit_prices = {}
        for tb in self.tb_data:
            tb_name = tb['投标信息']['投标人']
            if tb_name == selected_zbdw:
                # if tb['投标信息·'].get('投标人') == selected_zbdw:
                for sp in tb["单项工程"]:
                    for ue in sp['单位工程']:
                        for li in ue['分部清单']:
                            bid_unit_prices[li['编码']] = float(li[key])
                break

        for code, info in self.list_items.items():
            avg_price = mean(info['单价']) if info['单价'] else 0
            kzj_price = info['控制价单价']
            base_price = kzj_price * kzj_ratio + avg_price * (1 - kzj_ratio)
            bid_price = bid_unit_prices.get(code, 0)
            unit_diff = bid_price - base_price
            unit_rate = (unit_diff / base_price * 100) if base_price != 0 else 0
            shuliang = float(info['数量'])
            total_diff = unit_diff * shuliang
            if unit_diff < 0 and (condition == "或" and (abs(unit_rate) > pianchalv or abs(total_diff) > pianchazhi) or
                                  condition == "且" and abs(unit_rate) > pianchalv and abs(total_diff) > pianchazhi):
                table_data.append([
                    code, info['名称'], info['项目特征'], info['单位'], f"{shuliang:.2f}",
                    f"{kzj_price:.2f}", f"{avg_price:.2f}", f"{base_price:.2f}",
                    f"{bid_price:.2f}", f"{unit_diff:.2f}", f"{unit_rate:.2f}%", f"{total_diff:.2f}"
                ])

        if not table_data:
            table_data.append(["无偏低项"] + [""] * (len(self.diyu_table_titles) - 1))

        self.diyu_table.setRowCount(len(table_data))
        for row_idx, row_data in enumerate(table_data):
            for col_idx, value in enumerate(row_data):
                self.diyu_table.setItem(row_idx, col_idx, QTableWidgetItem(value))


class HtmlDelegate(QStyledItemDelegate):
    def paint(self, painter: QPainter, option: QStyleOptionViewItem, index):
        options = QStyleOptionViewItem(option)
        self.initStyleOption(options, index)
        painter.save()
        doc = QTextDocument()
        doc.setHtml(options.text)
        doc.setTextWidth(options.rect.width())
        painter.translate(options.rect.left(), options.rect.top())
        doc.drawContents(painter)
        painter.restore()

    def sizeHint(self, option: QStyleOptionViewItem, index):
        options = QStyleOptionViewItem(option)
        self.initStyleOption(options, index)
        doc = QTextDocument()
        doc.setHtml(options.text)
        doc.setTextWidth(options.rect.width())
        return doc.size().toSize()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    with open("./result.json", 'r', encoding="utf-8") as f:
        result = json.load(f)
    window = qingDanCompareWindow(result=result)
    window.show()
    app.exec()
