import datetime
import json
import os
import sys
import threading

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font as XLFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt, QUrl, QObject, Signal
from PySide6.QtGui import QAction, QKeySequence, QFont, QIcon, QDesktopServices
from PySide6.QtWidgets import QApplication, QDialog, QVBoxLayout, QHBoxLayout, \
    QListWidget, QStackedWidget, QWidget, QLabel, \
    QPushButton, QTextEdit, QAbstractItemView, QInputDialog, QSystemTrayIcon, QTableWidget, QTabWidget, \
    QFileDialog, QTableWidgetItem, QMessageBox, QLineEdit, QMainWindow, QProgressDialog

from about_page import AboutPage
from app_paths import resource_path
from read_data import read_file
from result import qingbiaoResult, zbfx_window
from statics import StaticSource
from updater import DownloadWindow
from utils import functions
from 对比清单差异性 import force_align_until_success


class WorkerSignals(QObject):
    """定义工作线程的信号"""
    progress = Signal(str, int)  # 任务名称, 进度值
    result = Signal(str, object)  # 任务名称, 结果
    error = Signal(str, str)  # 任务名称, 错误信息
    finished = Signal()


class QingbiaoWorker:
    """清标工作器类"""

    def __init__(self):
        self.signals = WorkerSignals()
        self._cancel_flag = False
        self._lock = threading.Lock()

    def cancel(self):
        """取消任务"""
        with self._lock:
            self._cancel_flag = True

    def is_cancelled(self):
        """检查是否已取消"""
        with self._lock:
            return self._cancel_flag


class main_window(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(StaticSource.get_software_name())
        self.resize(900, 700)
        self.old_value = ''

        # 初始化数据
        self.zb_data = {}
        self.kzj_data = {}
        self.tb_data = []
        self.pianchalv = 10.0

        # 窗口图标（Windows）
        self.setWindowIcon(QIcon(resource_path("icons/mac_icon.png")))

        # 系统托盘图标（Windows + macOS 通用）
        self.tray_icon = QSystemTrayIcon(self)

        # 优先用平台专用图标，不存在时回退通用
        icon_path = (
            "icons/mac_icon.icns" if sys.platform == "darwin" else
            "icons/win_icon.ico" if sys.platform.startswith("win") else
            "icons/app.png"  # Linux 或备选
        )

        tray_icon_qicon = QIcon(resource_path(icon_path))
        self.tray_icon.setIcon(tray_icon_qicon)
        self.tray_icon.setToolTip(StaticSource.get_software_name())
        self.tray_icon.show()

        self.create_menu()
        self.create_central_widget()
        self.update_btns_status()

        self.qingbiaoed = False

    def create_menu(self):
        menubar = self.menuBar()

        # 文件菜单（只保留数据保存/读取/清空/修改基本信息）
        file_menu = menubar.addMenu("文件")

        self.save_data_action = QAction("保存数据", self)
        self.save_data_action.setShortcut(QKeySequence("Ctrl+S"))
        self.save_data_action.triggered.connect(self.save_data)
        file_menu.addAction(self.save_data_action)

        self.read_data_action = QAction("读取数据", self)
        self.read_data_action.setShortcut(QKeySequence("Ctrl+O"))
        self.read_data_action.triggered.connect(self.read_data)
        file_menu.addAction(self.read_data_action)

        self.clear_data_action = QAction("清空数据", self)
        self.clear_data_action.setShortcut(QKeySequence("Ctrl+N"))
        self.clear_data_action.triggered.connect(self.clear_data)
        file_menu.addAction(self.clear_data_action)

        self.transcribe_data_action = QAction("数据转写", self)
        self.transcribe_data_action.triggered.connect(self.export_source_file_to_excel)
        file_menu.addAction(self.transcribe_data_action)

        file_menu.addSeparator()

        self.change_base_info_action = QAction("修改基本信息", self)
        self.change_base_info_action.triggered.connect(self.change_base_info)
        file_menu.addAction(self.change_base_info_action)

        # 数据导入菜单（新增）
        import_menu = menubar.addMenu("数据导入")

        self.import_xazb_action = QAction("导入招标清单文件", self)
        self.import_xazb_action.setShortcut(QKeySequence("Ctrl+I"))
        self.import_xazb_action.triggered.connect(self.import_xazb_file)
        import_menu.addAction(self.import_xazb_action)

        self.import_xaxj_action = QAction("导入招标控制价文件", self)
        self.import_xaxj_action.setShortcut(QKeySequence("Ctrl+Shift+I"))
        self.import_xaxj_action.triggered.connect(self.import_xaxj_file)
        import_menu.addAction(self.import_xaxj_action)

        import_menu.addSeparator()

        self.import_xatb_action = QAction("导入单个投标文件", self)
        self.import_xatb_action.setShortcut(QKeySequence("Ctrl+T"))
        self.import_xatb_action.triggered.connect(self.import_xatb_file)
        import_menu.addAction(self.import_xatb_action)

        self.import_xatb_mutil_action = QAction("导入多个投标文件", self)
        self.import_xatb_mutil_action.setShortcut(QKeySequence("Ctrl+Shift+T"))
        self.import_xatb_mutil_action.triggered.connect(self.import_xatb_mutil_file)
        import_menu.addAction(self.import_xatb_mutil_action)

        self.select_file_folder_action = QAction("选择投标文件夹导入所有", self)
        self.select_file_folder_action.setShortcut(QKeySequence("Ctrl+F"))
        self.select_file_folder_action.triggered.connect(self.select_file_folder)
        import_menu.addAction(self.select_file_folder_action)

        import_menu.addSeparator()

        self.delete_tb_data_action = QAction("删除当前单位", self)
        self.delete_tb_data_action.setShortcut(QKeySequence("Del"))
        self.delete_tb_data_action.triggered.connect(self.delete_tb_data)
        import_menu.addAction(self.delete_tb_data_action)

        # 清标菜单
        qingbiao_menu = menubar.addMenu("清标")

        self.deviation_action = QAction("设置单价偏差率...", self)
        self.deviation_action.triggered.connect(self.set_deviation)
        qingbiao_menu.addAction(self.deviation_action)

        self.include_shebei_action = QAction("单价含设备", self)
        self.include_shebei_action.setCheckable(True)
        self.include_shebei_action.setChecked(False)
        qingbiao_menu.addAction(self.include_shebei_action)

        self.start_qingbiao_action = QAction("开始清标", self)
        self.start_qingbiao_action.setShortcut(QKeySequence("Ctrl+G"))
        self.start_qingbiao_action.triggered.connect(self.start_qingbiao_jiude)
        qingbiao_menu.addAction(self.start_qingbiao_action)

        # self.start_qingbiao_action_jiude = QAction("开始清标(旧的)", self)
        # self.start_qingbiao_action_jiude.triggered.connect(self.start_qingbiao_jiude)
        # qingbiao_menu.addAction(self.start_qingbiao_action_jiude)

        # 中标菜单
        zb_menu = menubar.addMenu("中标")

        self.zbfx_action = QAction("中标分析", self)
        self.zbfx_action.setShortcut(QKeySequence("Ctrl+A"))
        self.zbfx_action.triggered.connect(self.zbfx)
        zb_menu.addAction(self.zbfx_action)

        # 帮助菜单
        help_menu = menubar.addMenu("帮助")

        self.help_action = QAction("帮助", self)
        self.help_action.setShortcut(QKeySequence("F1"))
        self.help_action.triggered.connect(self.show_help)
        help_menu.addAction(self.help_action)

        self.about_action = QAction("关于", self)
        self.about_action.triggered.connect(self.show_about)
        help_menu.addAction(self.about_action)

        self.check_update_action = QAction("检查更新", self)
        self.check_update_action.triggered.connect(self.check_update)
        help_menu.addAction(self.check_update_action)

    def msg_by_tray_icon(self, title, msg):
        information_icon = QIcon(resource_path("icons/win_icon.ico"))
        self.tray_icon.showMessage(
            title,
            msg,
            information_icon,
            2000
        )

    def show_about(self):
        about = AboutPage()
        about.exec()

    def check_update(self):
        current_version = StaticSource.get_current_version()
        latest_version = functions.get_latest_version()
        if not latest_version:
            QMessageBox.warning(self, "更新提示", "暂时无法获取最新版本信息，请稍后重试。")
            return

        result = functions.compare_version_numbers(current_version, latest_version)

        if result:
            reply = QMessageBox.question(self, "提示", "有新版本可用，是否前往下载？",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if reply == QMessageBox.StandardButton.Yes:
                # 统一使用下载文件夹，避免临时目录被自动删除的问题
                download_dir = os.path.join(os.path.expanduser("~"), "Downloads")

                # 确保下载目录存在
                os.makedirs(download_dir, exist_ok=True)

                print("下载目录：", download_dir)
                download_win = DownloadWindow(download_dir)
                download_win.exec()
            else:
                self.msg_by_tray_icon(title="更新提示", msg="已取消更新")
        else:
            message = f"当前版本：{current_version}，已是最新版本"
            self.msg_by_tray_icon(title="更新提示", msg=message)

    def change_base_info(self):
        self.base_info_data = functions.get_base_info_data()
        window = base_info(info=self.base_info_data)
        window.exec()
        self.base_info_data = functions.get_base_info_data()

    def create_central_widget(self):
        central_widget = QWidget()
        layout = QHBoxLayout(central_widget)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Tab标签
        self.tabs = QTabWidget()
        self.tab_1 = QTableWidget()
        self.tab_1_header = ["项目名称", "控制价金额"]
        self.tab_1.setColumnCount(len(self.tab_1_header))
        self.tab_1.setHorizontalHeaderLabels(self.tab_1_header)
        self.tab_2 = QTableWidget()
        self.tab_2_header = ["投标单位名称", "投标金额", "投标有效性"]
        self.tab_2.setColumnCount(len(self.tab_2_header))
        self.tab_2.setHorizontalHeaderLabels(self.tab_2_header)
        self.tab_2.cellClicked.connect(self.get_old_value)
        self.tab_2.itemChanged.connect(self.update_tb_data_unit)
        self.tabs.addTab(self.tab_1, "招标控制价")
        self.tabs.addTab(self.tab_2, "投标文件")

        layout.addWidget(self.tabs)
        self.setCentralWidget(central_widget)

    def set_deviation(self):
        value, ok = QInputDialog.getDouble(self, "设置偏差率", "单价偏差率（%）:", self.pianchalv, 0.0, 70.0, 2)
        if ok:
            self.pianchalv = value

    def update_btns_status(self):
        if self.zb_data:
            self.import_xazb_action.setText("招标清单已导入")
            self.import_xazb_action.setEnabled(False)
        else:
            self.import_xazb_action.setText("导入招标清单文件")
            self.import_xazb_action.setEnabled(True)
        if self.kzj_data:
            self.import_xaxj_action.setText("招标控制价已导入")
            self.import_xaxj_action.setEnabled(False)
        else:
            self.import_xaxj_action.setText("导入招标控制价文件")
            self.import_xaxj_action.setEnabled(True)

    def show_help(self):
        dialog = HelpDialog()
        dialog.exec()

    def import_xazb_file(self):
        print("导入招标清单文件")
        file, _ = QFileDialog.getOpenFileName(self, "选择招标清单文件", "", "XAZB Files (*.xazb *.XAZB)")
        if not file:
            return
        self.zb_data = read_file(file)
        self.update_btns_status()
        print("招标清单已导入")

    def import_xaxj_file(self):
        print("导入招标控制价文件")
        file, _ = QFileDialog.getOpenFileName(self, "选择招标控制价文件", "", "XAXJ Files (*.xaxj *.XAXJ)")
        if not file:
            return
        self.kzj_data = read_file(file)
        self.update_btns_status()
        self.update_zb_table()
        print("招标控制价已导入")

    def calculateTheBidDrawdownRate(self, tb_jine):
        tb_jine = float(tb_jine)
        kzj_jine = float(self.kzj_data['招标控制价信息']['控制价总价'])
        rate = (tb_jine / kzj_jine) * 100
        return rate

    def import_xatb_file(self):
        if not self.kzj_data:
            QMessageBox.information(self, "提示", "无控制价数据，请先导入控制价文件")
            return

        print("导入投标文件")
        file, _ = QFileDialog.getOpenFileName(self, "选择投标文件", "", "XATB Files (*.xatb *.XATB)")
        if not file:
            return

        # 读取原始数据
        tb_data = read_file(file)
        tb_name = tb_data['投标信息']['投标人']

        # 关键：结构对齐（使用最新版 force_align_until_success + 跳过功能）
        aligned_kzj, aligned_tb, skipped = force_align_until_success(
            self.kzj_data,  # 控制价（基准）
            tb_data,  # 投标文件（待对齐）
            name1="控制价文件",
            name2=tb_name
        )

        if skipped:
            print("用户跳过此投标文件导入")
            return

        if not aligned_kzj or not aligned_tb:
            print("对齐失败，取消导入")
            return

        # 更新内存中的控制价（因为用户可能调整了顺序）
        self.kzj_data = aligned_kzj

        # 计算下浮率
        tb_data_rate = self.calculateTheBidDrawdownRate(tb_jine=aligned_tb['投标信息']['投标总价'])
        aligned_tb['投标信息']['下浮率'] = tb_data_rate

        # 成功导入
        self.tb_data.append(aligned_tb)
        self.update_btns_status()
        self.update_tb_table()

        print(f"投标文件已成功导入并对齐：{os.path.basename(file)}")

    def import_xatb_mutil_file(self):
        print("批量导入xatb文件")
        if not self.kzj_data:
            QMessageBox.information(self, "提示", "无控制价数据，请先导入控制价文件")
            return
        files, _ = QFileDialog.getOpenFileNames(self, "选择投标文件", "", "XATB Files (*.xatb *.XATB)")
        if not files:
            return
        for file in files:
            print(f"正在导入{file}")
            tb_data = read_file(file)
            tb_name = tb_data['投标信息']['投标人']
            aligned_kzj, aligned_tb, skipped = force_align_until_success(
                self.kzj_data,  # 控制价（基准）
                tb_data,  # 投标文件（待对齐）
                name1="控制价文件",
                name2=tb_name
            )
            if skipped:
                print("用户跳过此投标文件导入")
                continue
            if not aligned_kzj or not aligned_tb:
                print("对齐失败，取消导入")
                continue

            # 更新内存中的控制价（因为用户可能调整了顺序）
            self.kzj_data = aligned_kzj
            # 计算下浮率
            tb_data_rate = self.calculateTheBidDrawdownRate(tb_jine=aligned_tb['投标信息']['投标总价'])
            aligned_tb['投标信息']['下浮率'] = tb_data_rate

            self.tb_data.append(aligned_tb)
            print(f"{file}导入完成")
        self.update_tb_table()
        self.msg_by_tray_icon(title="导入完成", msg=f"所有xatb文件导入完成，共导入{len(files)}个投标文件")

    def select_file_folder(self):
        print("选择xatb文件夹")
        if not self.kzj_data:
            QMessageBox.information(self, "提示", "无控制价数据，请先导入控制价文件")
            return
        folder = QFileDialog.getExistingDirectory(self, "选择xatb文件夹")
        if not folder:
            return

        # 统计所有待导入的xatb文件数量
        xatb_files = []
        for root, dirs, files in os.walk(folder):
            for file in files:
                if file.endswith(".xatb") or file.endswith(".XATB"):
                    xatb_files.append(os.path.join(root, file))
        total_files = len(xatb_files)
        if total_files == 0:
            QMessageBox.information(self, "提示", "未找到xatb文件")
            return

        # 创建进度条对话框
        progress = QProgressDialog("正在导入投标文件...", "取消", 0, total_files, self)
        progress.setWindowTitle("导入进度")
        progress.setWindowModality(Qt.WindowModality.ApplicationModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)

        files_count = 0
        for idx, file_path in enumerate(xatb_files, 1):
            if progress.wasCanceled():
                break
            file_name = os.path.basename(file_path)
            progress.setLabelText(f"正在导入【{file_name}】 ({idx}/{total_files})")
            QApplication.processEvents()
            print(f"正在导入【{file_name}】")
            tb_data = read_file(file_path)
            tb_name = tb_data['投标信息']['投标人']
            aligned_kzj, aligned_tb, skipped = force_align_until_success(
                self.kzj_data,  # 控制价（基准）
                tb_data,  # 投标文件（待对齐）
                name1="控制价文件",
                name2=tb_name
            )
            if skipped:
                print("用户跳过此投标文件导入")
                continue

            if not aligned_kzj or not aligned_tb:
                print("对齐失败，取消导入")
                continue

            self.kzj_data = aligned_kzj
            # 计算下浮率
            tb_data_rate = self.calculateTheBidDrawdownRate(tb_jine=aligned_tb['投标信息']['投标总价'])
            aligned_tb['投标信息']['下浮率'] = tb_data_rate

            self.tb_data.append(aligned_tb)
            files_count += 1
            print(f"【{file_name}】导入完成")
            progress.setValue(idx)

        progress.close()
        self.update_tb_table()
        print(f"所有xatb文件导入完成，共导入{files_count}个投标文件")
        QMessageBox.information(self, "提示", f"所有xatb文件导入完成，共导入{files_count}个投标文件")

    def update_zb_table(self):
        if not self.kzj_data:
            self.tab_1.clear()  # 清空表格
            return
        self.tab_1.setRowCount(1)
        project_name = self.kzj_data['项目信息']['项目名称']
        project_jine = self.kzj_data['招标控制价信息']['控制价总价']
        self.tab_1.setItem(0, 0, QTableWidgetItem(str(project_name)))
        self.tab_1.setItem(0, 1, QTableWidgetItem(str(project_jine)))

    def update_tb_table(self):
        self.tab_2.setRowCount(len(self.tb_data))
        for i, tb_data in enumerate(self.tb_data):
            tender_unit_name = tb_data['投标信息']['投标人']
            tender_jine = tb_data['投标信息']['投标总价']
            is_checked = tb_data['投标信息'].get("有效性", True)
            self.tab_2.setItem(i, 0, QTableWidgetItem(str(tender_unit_name)))
            self.tab_2.setItem(i, 1, QTableWidgetItem(str(tender_jine)))
            # self.tab_2.setItem(i, 2, QTableWidgetItem(f"{tb_rate:.2f}%"))
            checkbox_item = QTableWidgetItem()
            checkbox_item.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            if is_checked:
                checkbox_item.setCheckState(Qt.CheckState.Checked)
            else:
                checkbox_item.setCheckState(Qt.CheckState.Unchecked)
            self.tab_2.setItem(i, 2, checkbox_item)
            self.tab_2.setColumnWidth(2, 50)
            # 在这增加个复选框，每行单元格都增加一个复选框

    def get_old_value(self, row, column):
        if column != 0:
            return
        try:
            self.old_value = self.tab_2.item(row, column).text() or ""
        except:
            self.old_value = ""

    # 写一个修改投标文件的投标单位的方法，随时监听表格变化，并替换、修改对应投标文件中的
    def update_tb_data_unit(self, item):
        # 监听第一列的变换
        if item.column() == 0:
            new_name = item.text()
            for tb_data in self.tb_data:
                if tb_data['投标信息']['投标人'] == self.old_value:
                    tb_data['投标信息']['投标人'] = new_name  # 赋新值
                    break
        # 监听第四列变化，复选框的变化情况
        elif item.column() == 3:
            row = item.row()
            company_name_item = self.tab_2.item(row, 0)
            if not company_name_item:
                return
            company_name = company_name_item.text()
            is_checked = item.checkState() == Qt.CheckState.Checked
            self.update_tb_is_ok(company_name, is_checked)
        else:
            return

    def update_tb_is_ok(self, company_name: str, is_checked: bool):
        for tb in self.tb_data:
            if tb['投标信息']['投标人'] == company_name:
                tb['投标信息']['有效性'] = is_checked
                break

    def delete_tb_data(self):
        if self.tabs.currentIndex() == 0:
            return
        if not self.tab_2.currentItem():
            return
        row = self.tab_2.currentRow()
        company_name = self.tab_2.item(row, 0).text()
        self.tb_data = [data for data in self.tb_data if data['投标信息']['投标人'] != company_name]
        reply = QMessageBox.question(self, '确认删除', f'确定要删除投标单位 "{company_name}" 吗？',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No:
            return
        self.msg_by_tray_icon(title="删除投标文件", msg=f"投标文件【{company_name}】已删除")
        self.update_tb_table()

    def save_data(self):
        if not self.zb_data or not self.kzj_data or not self.tb_data:
            QMessageBox.warning(self, "提示", "请先导入招标清单、招标控制价、投标文件")
            return
        all_data = {
            "招标清单": self.zb_data,
            "招标控制价": self.kzj_data,
            "投标文件": self.tb_data,
        }
        time_now = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        file, _ = QFileDialog.getSaveFileName(self, "保存数据", f"清标数据{time_now}.json", "JSON Files (*.json)")
        if not file:
            return
        with open(file, "w", encoding="utf-8") as f:
            json.dump(all_data, f, ensure_ascii=False, indent=4)

        print(f"数据已保存为JSON文件: {file}")

    def read_data(self):
        file, _ = QFileDialog.getOpenFileName(self, "选择数据文件", "", "JSON Files (*.json)")
        if not file:
            return
        with open(file, "r", encoding="utf-8") as f:
            data = json.load(f)

        self.zb_data = data.get("招标清单", {})
        self.kzj_data = data.get("招标控制价", {})
        self.tb_data = data.get("投标文件", [])
        self.update_btns_status()
        self.update_tb_table()
        self.update_zb_table()
        print(f"数据已读取，已读取{len(self.tb_data)}条投标数据")

    def clear_data(self):
        reply = QMessageBox.question(self, "确认清空", "确定要清空所有数据吗？此操作不可撤销。",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.zb_data = {}
            self.kzj_data = {}
            self.tb_data = []
            self.update_btns_status()
            self.update_tb_table()
            self.update_zb_table()
            self.tray_icon.showMessage("提示", "数据已清空", QSystemTrayIcon.MessageIcon.Warning)
            self.msg_by_tray_icon(title="清空数据", msg="数据已清空")
            print("数据已清空")
        else:
            print("取消清空数据")

    def export_source_file_to_excel(self):
        source_file, _ = QFileDialog.getOpenFileName(
            self,
            "选择需要转写的数据文件",
            "",
            "XML Files (*.xatb *.xazb *.xaxj *.XATB *.XAZB *.XAXJ)"
        )
        if not source_file:
            return

        try:
            parsed_data = read_file(source_file)
        except Exception as exc:
            QMessageBox.critical(self, "转写失败", f"文件解析失败：{exc}")
            return

        default_name = f"{os.path.splitext(os.path.basename(source_file))[0]}.xlsx"
        save_file, _ = QFileDialog.getSaveFileName(
            self,
            "保存转写结果",
            default_name,
            "Excel Files (*.xlsx)"
        )
        if not save_file:
            return

        if not save_file.lower().endswith(".xlsx"):
            save_file += ".xlsx"

        try:
            workbook = self.build_transcribe_workbook(parsed_data, source_file)
            workbook.save(save_file)
            self.msg_by_tray_icon("数据转写", f"已导出 Excel：{save_file}")
            QMessageBox.information(self, "导出成功", f"数据已转写为 Excel：\n{save_file}")
        except Exception as exc:
            QMessageBox.critical(self, "导出失败", f"生成 Excel 失败：{exc}")

    def build_transcribe_workbook(self, data: dict, source_file: str) -> Workbook:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        self.write_summary_sheet(workbook, data, source_file)
        self.write_info_sheet(workbook, "项目信息", data.get("项目信息", {}))
        self.write_info_sheet(workbook, "招标信息", data.get("招标信息", {}))
        self.write_info_sheet(workbook, "投标信息", data.get("投标信息", {}))
        self.write_info_sheet(workbook, "控制价信息", data.get("招标控制价信息", {}))

        fenbu_rows = []
        cuoshi_rows = []
        other_rows = []

        for project in data.get("单项工程", []):
            project_name = project.get("名称", "")
            for unit in project.get("单位工程", []):
                unit_name = unit.get("名称", "")
                for item in unit.get("分部清单", []):
                    fenbu_rows.append({
                        "单项工程": project_name,
                        "单位工程": unit_name,
                        "编码": item.get("编码", ""),
                        "名称": item.get("名称", ""),
                        "项目特征": item.get("项目特征", ""),
                        "单位": item.get("单位", ""),
                        "数量": item.get("数量", ""),
                        "综合单价": item.get("综合单价", ""),
                        "综合合价": item.get("综合合价", ""),
                        "设备单价": item.get("设备单价", ""),
                        "综合单价_含设备": item.get("综合单价_含设备", ""),
                    })

                for item in unit.get("措施清单", []):
                    cuoshi_rows.append({
                        "单项工程": project_name,
                        "单位工程": unit_name,
                        "编码": item.get("编码", ""),
                        "名称": item.get("名称", ""),
                        "项目特征": item.get("项目特征", ""),
                        "单位": item.get("单位", ""),
                        "数量": item.get("数量", ""),
                        "综合单价": item.get("综合单价", ""),
                        "综合合价": item.get("综合合价", ""),
                    })

                other_items = unit.get("其他项目", {})
                for category, rows in other_items.items():
                    for item in rows:
                        row = {
                            "单项工程": project_name,
                            "单位工程": unit_name,
                            "其他项目类型": category,
                        }
                        row.update(item)
                        other_rows.append(row)

        self.write_table_sheet(
            workbook,
            "分部清单",
            [
                "单项工程", "单位工程", "编码", "名称", "项目特征",
                "单位", "数量", "综合单价", "综合合价", "设备单价", "综合单价_含设备"
            ],
            fenbu_rows
        )
        self.write_table_sheet(
            workbook,
            "措施清单",
            ["单项工程", "单位工程", "编码", "名称", "项目特征", "单位", "数量", "综合单价", "综合合价"],
            cuoshi_rows
        )
        other_headers = ["单项工程", "单位工程", "其他项目类型"]
        for row in other_rows:
            for key in row.keys():
                if key not in other_headers:
                    other_headers.append(key)
        self.write_table_sheet(workbook, "其他项目", other_headers, other_rows)

        return workbook

    def write_summary_sheet(self, workbook: Workbook, data: dict, source_file: str):
        summary_rows = [
            ("源文件名称", os.path.basename(source_file)),
            ("源文件路径", source_file),
            ("文件类型", os.path.splitext(source_file)[1].lower() or "未知"),
            ("导出时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("单项工程数量", len(data.get("单项工程", []))),
            ("分部清单条数", self.count_data_items(data, "分部清单")),
            ("措施清单条数", self.count_data_items(data, "措施清单")),
            ("其他项目条数", self.count_other_items(data)),
        ]
        sheet = workbook.create_sheet("文件概览")
        sheet.append(["项目", "内容"])
        for row in summary_rows:
            sheet.append(list(row))
        self.style_info_sheet(sheet, title="数据转写概览")

    def write_info_sheet(self, workbook: Workbook, sheet_name: str, info: dict):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(["字段", "值"])
        if info:
            for key, value in info.items():
                sheet.append([key, value])
        else:
            sheet.append(["说明", "该部分无数据"])
        self.style_info_sheet(sheet, title=sheet_name)

    def write_table_sheet(self, workbook: Workbook, sheet_name: str, headers: list[str], rows: list[dict]):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        if rows:
            for row in rows:
                sheet.append([row.get(header, "") for header in headers])
        else:
            sheet.append(["暂无数据"] + [""] * (len(headers) - 1))
        self.style_table_sheet(sheet, title=sheet_name, wrap_columns={"项目特征", "值", "内容", "源文件路径"})

    def style_info_sheet(self, sheet, title: str):
        self.insert_sheet_title(sheet, title)
        self.apply_sheet_style(sheet, wrap_columns={"值", "内容"}, key_columns={"字段", "项目"})

    def style_table_sheet(self, sheet, title: str, wrap_columns: set[str] | None = None):
        self.insert_sheet_title(sheet, title)
        self.apply_sheet_style(sheet, wrap_columns=wrap_columns or set())

    def insert_sheet_title(self, sheet, title: str):
        sheet.insert_rows(1)
        sheet["A1"] = title
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, sheet.max_column))
        cell = sheet["A1"]
        cell.font = XLFont(size=14, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 24

    def apply_sheet_style(self, sheet, wrap_columns: set[str], key_columns: set[str] | None = None):
        header_row = 2
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        stripe_fill = PatternFill("solid", fgColor="F7FBFF")

        for cell in sheet[header_row]:
            cell.font = XLFont(bold=True, color="1F1F1F")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        header_names = [sheet.cell(header_row, col).value for col in range(1, sheet.max_column + 1)]
        wrap_index = {idx + 1 for idx, name in enumerate(header_names) if name in wrap_columns}
        key_index = {idx + 1 for idx, name in enumerate(header_names) if key_columns and name in key_columns}

        for row in range(header_row + 1, sheet.max_row + 1):
            row_fill = stripe_fill if row % 2 == 1 else None
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                if col in wrap_index:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                elif col in key_index:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.font = XLFont(bold=True)
                else:
                    cell.alignment = Alignment(vertical="center")

        sheet.freeze_panes = "A3"
        for idx, header in enumerate(header_names, start=1):
            max_length = len(str(header)) if header else 0
            for row in range(header_row + 1, sheet.max_row + 1):
                value = sheet.cell(row, idx).value
                if value is None:
                    continue
                value_len = len(str(value).replace("\n", ""))
                if idx in wrap_index:
                    value_len = min(value_len, 30)
                max_length = max(max_length, value_len)
            sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 40)

        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"

    def count_data_items(self, data: dict, key: str) -> int:
        count = 0
        for project in data.get("单项工程", []):
            for unit in project.get("单位工程", []):
                count += len(unit.get(key, []))
        return count

    def count_other_items(self, data: dict) -> int:
        count = 0
        for project in data.get("单项工程", []):
            for unit in project.get("单位工程", []):
                for rows in unit.get("其他项目", {}).values():
                    count += len(rows)
        return count

    def open_save_path(self):
        if hasattr(self, "save_path") and self.save_path:
            QDesktopServices.openUrl(QUrl.fromLocalFile(self.save_path))

    def start_qingbiao_jiude(self):
        """
        执行清标计算，使用多线程并行计算各项结果（含实时进度显示）
        """
        # 检查数据是否已导入
        if not self.kzj_data or not self.zb_data or not self.tb_data:
            QMessageBox.warning(self, "错误", "请先导入数据")
            return

        # 创建进度窗口
        self.progress_window = QDialog(self)
        self.progress_window.setWindowTitle("清标计算中，请稍候...")
        self.progress_window.resize(700, 500)
        layout = QVBoxLayout()
        self.progress_text = QTextEdit()
        self.progress_text.setReadOnly(True)
        layout.addWidget(self.progress_text)
        self.progress_window.setLayout(layout)
        self.progress_window.show()
        QApplication.processEvents()

        def log(msg):
            self.progress_text.append(msg)
            QApplication.processEvents()

        # 获取是否计算设备单价
        is_include_shebei = self.include_shebei_action.isChecked()
        log(f"是否包含设备费计算：{is_include_shebei}")
        log("开始清标计算...")

        start_time = datetime.datetime.now()
        log(f"开始时间：{start_time.strftime('%Y-%m-%d %H:%M:%S')}")

        try:
            # 总价限价结果
            log("正在计算总价限价...")
            zongjia_jieguo = functions.get_zongjia_jieguo(self.kzj_data, self.tb_data)
            log("总价限价计算完成")

            # 负值、零值结果
            log("正在检查负值与零值...")
            fuzhi_jieguo, zero_jieguo = functions.get_fuzhi_and_zero_jieguo(tb_datas=self.tb_data)
            log("负值、零值检查完成")

            # 正负10%偏差（已只保留_10_jieguo）
            log("正在计算正负偏差率项目...")
            _10_jieguo = functions.get_10_result(self.kzj_data, self.tb_data, self.pianchalv, is_include_shebei)
            log("偏差率计算完成")

            # 获取清单对比结果（若仍需可保留）
            log("正在进行清单一致性对比...")
            print("正在获取清单对比结果...")
            new_qingdan_jieguo = functions.get_qingdan_result(self.kzj_data, self.tb_data)
            log("清单对比完成")
            print("清单对比结果：", new_qingdan_jieguo)

            # 硬件加密锁检查
            log("正在检查硬件加密锁...")
            jiamisuo_jieguo = functions.get_jiamisuo_jieguo(tb_data=self.tb_data)
            log("加密锁检查完成")

            end_time = datetime.datetime.now()
            cost = end_time - start_time
            log(f"清标全部完成，耗时 {cost.total_seconds():.2f} 秒")

            # 关闭进度窗，打开结果窗
            self.progress_window.accept()
            win = qingbiaoResult(
                zongjia_jieguo,
                fuzhi_jieguo,
                zero_jieguo,
                new_qingdan_jieguo,
                jiamisuo_jieguo,
                _10_jieguo,
                self.pianchalv,
                self.tray_icon
            )
            win.exec()
            self.qingbiaoed = True

        except Exception as e:
            log(f"清标过程出错：{str(e)}")
            QMessageBox.critical(self, "清标失败", str(e))
            self.progress_window.accept()

    def gener_report(self):
        if not self.qingbiaoed:
            QMessageBox.warning(self, "错误", "请先执行清标")


    def zbfx(self):
        print("正在执行招标分析")
        if not self.kzj_data or not self.zb_data or not self.tb_data:
            QMessageBox.warning(self, "错误", "请先导入招标清单、招标控制价和投标文件")
            return
        zbfx_window_show = zbfx_window(zb_data=self.zb_data, tb_data=self.tb_data, kzj_data=self.kzj_data)
        zbfx_window_show.exec()



class base_info(QDialog):
    def __init__(self, info=None):
        super().__init__()
        self.setWindowTitle("基本信息录入")
        self.setWindowIcon(QIcon(resource_path("icons/base_info.png")))
        self.info = info
        self.init()
        self.update_info()

    def init(self):
        layout = QVBoxLayout()

        jsdw_layout = QHBoxLayout()
        jsdw_label = QLabel("建设单位：")
        self.jsdw_input = QLineEdit()
        jsdw_layout.addWidget(jsdw_label)
        jsdw_layout.addWidget(self.jsdw_input)

        bzdw_layout = QHBoxLayout()
        bzdw_label = QLabel("编制单位：")
        self.bzdw_input = QLineEdit()
        bzdw_layout.addWidget(bzdw_label)
        bzdw_layout.addWidget(self.bzdw_input)

        self.confirm_btn = QPushButton("确定")
        self.confirm_btn.clicked.connect(self.confirm)
        self.confirm_btn.setShortcut("Enter")

        layout.addLayout(jsdw_layout)
        layout.addLayout(bzdw_layout)
        layout.addWidget(self.confirm_btn)
        self.setLayout(layout)

    def update_info(self):
        if self.info:
            self.jsdw_input.setText(self.info.get("建设单位", ""))
            self.bzdw_input.setText(self.info.get("编制单位", ""))

    def confirm(self):
        jsdw = self.jsdw_input.text()
        bzdw = self.bzdw_input.text()
        if jsdw and bzdw:
            base_info_data = {
                "建设单位": jsdw,
                "编制单位": bzdw
            }
            with open(functions.get_base_info_data(), "w", encoding="utf-8") as file:
                json.dump(base_info_data, file, ensure_ascii=False, indent=4)
            print("基本信息已保存")
            self.close()


class HelpDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("帮助")
        self.setMinimumSize(800, 600)
        self.setWindowIcon(QIcon(resource_path("icons/help.png")))
        self.init_ui()

    def init_ui(self):
        main_layout = QHBoxLayout()

        # Left drawer (menu)
        self.menu_list = QListWidget()
        self.menu_list.addItems(["使用说明", "快捷键说明", "联系我"])
        self.menu_list.currentRowChanged.connect(self.switch_content)

        # Right content (stacked)
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.addWidget(self.create_usage_page())
        self.stacked_widget.addWidget(self.create_shortcuts_page())
        self.stacked_widget.addWidget(self.create_contact_page())

        main_layout.addWidget(self.menu_list, 1)
        main_layout.addWidget(self.stacked_widget, 3)
        self.setLayout(main_layout)

    def create_usage_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        title = QLabel("使用说明")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        info = QTextEdit()
        info.setReadOnly(True)
        info.setFont(QFont("Arial", 13))
        usage_text = """
            1. 数据导入：
               - 点击“导入招标清单文件”导入招标工程量清单（.xazb）。
               - 点击“导入招标控制价文件”导入控制价（.xaxj）。
               - 点击“导入单个投标文件”或“导入多个投标文件”导入投标文件（.xatb），或选择文件夹批量导入。
               - 导入后，可在“投标文件”标签查看并编辑投标单位名称或有效性（复选框）。
            
            2. 数据管理：
               - “保存数据”：保存当前数据为JSON文件（Ctrl+S）。
               - “读取数据”：从JSON文件加载数据（Ctrl+O）。
               - “清空数据”：清除所有导入数据。
            
            3. 清标操作：
               - 在“清标”组设置“单价偏差率”（默认10%）。
               - 增加可选“含设备单价”分析（设备单价不含税）
               - 点击“开始清标”（Ctrl+G），展示结果，表格可导出。
            
            4. 中标分析：
               - 点击“中标分析”打开分析窗口。
               - 设置控制价比例（默认40%）、选择中标单位、偏差率阈值。
               - 点击“开始分析”查看单项对比、偏高/低项表格。
               - 点击“生成报告”导出Excel报告。
            
            5. 其他：
               - 软件已开源免费，可直接使用全部核心功能。
        """
        info.setText(usage_text)
        layout.addWidget(info)

        return page

    def create_shortcuts_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        title = QLabel("快捷键设置")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # Note on platform differences
        note = QLabel("注：macOS上Ctrl键对应⌘Cmd键；Windows/Linux使用Ctrl键。")
        note.setFont(QFont("Arial", 12))
        note.setStyleSheet("color: red;")
        layout.addWidget(note)

        # Table 1: Main Window - Data Management
        layout.addWidget(QLabel("主窗口 - 数据管理"))
        table1 = QTableWidget()
        table1.setRowCount(4)
        table1.setColumnCount(2)
        table1.setHorizontalHeaderLabels(["按钮名称", "快捷键"])
        table1.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        shortcuts1 = [
            ("保存数据", "Ctrl+S"),
            ("读取数据", "Ctrl+O"),
            ("清空数据", "Ctrl+N"),
            ("开始清标", "Ctrl+G")
        ]
        for row, (action, key) in enumerate(shortcuts1):
            table1.setItem(row, 0, QTableWidgetItem(action))
            table1.setItem(row, 1, QTableWidgetItem(key))
        table1.resizeColumnsToContents()
        layout.addWidget(table1)

        # Table 2: Main Window - File Import
        layout.addWidget(QLabel("主窗口 - 文件导入"))
        table2 = QTableWidget()
        table2.setRowCount(7)
        table2.setColumnCount(2)
        table2.setHorizontalHeaderLabels(["按钮名称", "快捷键"])
        table2.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        shortcuts2 = [
            ("导入招标清单文件", "Ctrl+I"),
            ("导入招标控制价文件", "Ctrl+Shift+I"),
            ("导入单个投标文件", "Ctrl+T"),
            ("导入多个投标文件", "Ctrl+Shift+T"),
            ("选择投标文件夹导入所有", "Ctrl+F"),
            ("删除当前单位", "Del"),
            ("帮助", "F1")
        ]
        for row, (action, key) in enumerate(shortcuts2):
            table2.setItem(row, 0, QTableWidgetItem(action))
            table2.setItem(row, 1, QTableWidgetItem(key))
        table2.resizeColumnsToContents()
        layout.addWidget(table2)

        # Table 3: Analysis Window
        layout.addWidget(QLabel("中标分析窗口"))
        table3 = QTableWidget()
        table3.setRowCount(2)
        table3.setColumnCount(2)
        table3.setHorizontalHeaderLabels(["按钮名称", "快捷键"])
        table3.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        shortcuts3 = [
            ("开始分析", "Ctrl+B"),
            ("生成报告", "Ctrl+R")
        ]
        for row, (action, key) in enumerate(shortcuts3):
            table3.setItem(row, 0, QTableWidgetItem(action))
            table3.setItem(row, 1, QTableWidgetItem(key))
        table3.resizeColumnsToContents()
        layout.addWidget(table3)

        # Table 4: Dialogs
        layout.addWidget(QLabel("对话框"))
        table4 = QTableWidget()
        table4.setRowCount(2)
        table4.setColumnCount(2)
        table4.setHorizontalHeaderLabels(["按钮名称", "快捷键"])
        table4.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        shortcuts4 = [
            ("基本信息确定", "Enter"),
            ("打开保存位置", "Ctrl+H")
        ]
        for row, (action, key) in enumerate(shortcuts4):
            table4.setItem(row, 0, QTableWidgetItem(action))
            table4.setItem(row, 1, QTableWidgetItem(key))
        table4.resizeColumnsToContents()
        layout.addWidget(table4)

        return page

    def create_contact_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)

        title = QLabel("联系我")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        info = QTextEdit()
        info.setReadOnly(True)
        info.setFont(QFont("Arial", 13))
        info.setFixedHeight(100)  # 限制高度为100像素，大约5行文本
        info.setText("邮箱: wl201808087@icloud.com\n电话: 176 0040 9422\n微信: W851172873")

        layout.addWidget(info)
        layout.addStretch()

        return page

    def switch_content(self, index):
        self.stacked_widget.setCurrentIndex(index)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path("icons/mac_icon.png")))  # 通用ico
    window = main_window()
    window.show()
    sys.exit(app.exec())
